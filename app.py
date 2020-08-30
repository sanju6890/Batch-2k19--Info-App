from tkinter import *
from tkinter import messagebox
from openpyxl import *
file = load_workbook('ece.xlsx')
s = file['Sheet1']

root=Tk()
root.title("Student Info")
root.geometry("550x300")
root.configure(bg='cyan2')

def find():
    try:
        flag=0
        roll_no = int(box.get())
        row = s.max_row
        col = s.max_column
        for row_no in range(3, row+1):
            if  s.cell(row_no,1).value == roll_no:
                flag=1
                data=f"Student's Name: {s.cell(row_no,2).value.upper()}\nRoll no: {s.cell(row_no,1).value}\nBranch: {s.cell(row_no,4).value}\nE-mail id: {s.cell(row_no,3).value}"
                #result.config(text=data)
                messagebox.showinfo('Student Info',data)
                break
            row_no = row_no+1
        if flag == 0:
            messagebox.showwarning('Student Info','Record Not Found')
    except ValueError:
        messagebox.showerror('Student Info','Incorrect Input')


header = Label(root, text='Electronics and Communication Engineering 2k19',bg="cyan4",fg="white",font=("Times", "16", "bold italic"),borderwidth=5, relief=SUNKEN)
header.pack(pady=10)

label = Label(root, text='Enter Roll No',bg='black',fg='white',font=("Times", "14", "bold"),borderwidth=5,relief=SUNKEN)
label.pack(pady=10)

box = Entry(root,font=("Times", "12", "bold"),borderwidth=5)
box.pack()

button = Button(root, text='Search',bg="black",fg="white",font=("Times", "12", "bold"),borderwidth=5,command=find)
button.pack(pady=10)

# result = Label(root,text='',font=("Times", "14", "italic"),borderwidth=5)
# result.pack(padx=5)

label = Label(root, text='SANJAY KUMAR (C) 2020',bg='cyan2',font=("Times", "12", "bold"),borderwidth=5)
label.pack(pady=40)

root.mainloop()